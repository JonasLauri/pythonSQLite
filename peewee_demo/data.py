from openpyxl import load_workbook

# load excel file and make a sheet
workbook = load_workbook(filename="users_data.xlsx")
sheet = workbook.active

# iteration through the whole dataset and store in the list
users_info = []
for row in sheet.iter_rows(min_row=1,
                           min_col=1,
                           max_col=2,
                           values_only=True):
    users_info.append(row)

if __name__ == '__main__':
    print(users_info)